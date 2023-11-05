from django.shortcuts import render
from consulta.models import Livros, Estoque
from django.views.generic.list import ListView
from openpyxl import load_workbook


def index(request):
    if "GET" == request.method:
        return render(request,'index.html')
    if "POST" == request.method:
        filename = request.FILES['excel_file']
        wb = load_workbook(request.FILES['excel_file'])
        ws = wb.active
        dict_isbn = {}
        for row in ws['a']:
            if ws[f'A{row.row}'].value == 'ISBN':
                pass

            else:
                isbn = ws[f'A{row.row}'].value
                if str(isbn) in dict_isbn:
                    dict_isbn[f'{isbn}'] = int(dict_isbn[f'{isbn}']) + int(ws[f'D{row.row}'].value)

                else:
                    dict_isbn[f'{isbn}'] = int(ws[f'D{row.row}'].value)
            for item in dict_isbn:
                isbn = item
                qtd = dict_isbn[item]
                queryisbn = Livros.objects.get(isbn1=isbn)
                queryestoque = Estoque.objects.filter(nbook = queryisbn.nbook).first()
                estoque = queryestoque.disp
                estoque_ff = queryestoque.disp_ff
                if estoque >= qtd:
                    ws[f'B{row.row}'] = 'Circulo'

                elif estoque < qtd > 0:
                    estoque = estoque_ff + estoque
                    ws[f'B{row.row}'] = 'Catavento'
                else:
                    ws[f'B{row.row}'] = ''
            wb.save(f'{filename}')

        return render(request, 'index.html')


class SearchView(ListView):
    model = Livros
    template_name = 'search.html'
    context_object_name = 'result'

    # def get_context_data(self, *, object_list=None, **kwargs):
    #     result = self.queryset
    #     context = {
    #         'stock': self.stock,
    #         'result': result
    #     }
    #     return context
    def get_queryset(self):
       result = super(SearchView, self).get_queryset()
       query = self.request.GET.get('search')

       if query:
          postresult = Livros.objects.get(isbn1=query)
          #poststock = Estoque.objects.get(nbook=postresult.nbook)
          #self.estoquetotal = int(poststock.disp)
          result = postresult
          stock = self.get_estoque(postresult.nbook)
          result = postresult
          self.stock = stock
          self.result = result
       else:
           result = None
           self.result = result
           self.stock = None
       return (self.result, self.stock)

    def get_estoque(self, nbook):
        poststock = Estoque.objects.filter(nbook=nbook).first()
        self.poststock = poststock.disp + poststock.disp_ff
        return self.poststock